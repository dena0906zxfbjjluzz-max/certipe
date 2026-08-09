"""
CertiPE en Streamlit — misma cara que la web local (colores y layout).
"""

from __future__ import annotations

import csv
import io
import os
from datetime import datetime
from urllib.parse import quote

import streamlit as st

from app import certificates, crypto
from app import pdf_cert

APP_VERSION = "1.2 · look web local"

st.set_page_config(
    page_title="CertiPE · Certificados",
    page_icon="📜",
    layout="centered",
    initial_sidebar_state="collapsed",
)


def _secret_get(*paths: str) -> str | None:
    try:
        cur = st.secrets
        for p in paths:
            cur = cur[p]
        val = str(cur).strip()
        return val or None
    except Exception:
        return None


def bootstrap_secrets() -> None:
    seed = (
        _secret_get("LLAVE_PRIVADA")
        or _secret_get("credenciales", "LLAVE_PRIVADA")
        or os.environ.get("LLAVE_PRIVADA")
    )
    if seed:
        try:
            crypto.set_seed_hex(seed)
        except Exception as e:  # noqa: BLE001
            st.session_state["seed_error"] = str(e)

    nombre = (
        _secret_get("nombre_institucion")
        or _secret_get("credenciales", "nombre_institucion")
        or _secret_get("NOMBRE_INSTITUCION")
    )
    if nombre:
        os.environ["INSTITUTION_NAME"] = nombre

    inst_id = _secret_get("institution_id") or _secret_get("credenciales", "institution_id")
    if inst_id:
        os.environ["INSTITUTION_ID"] = inst_id


def cargar_credenciales() -> tuple[str | None, str | None, str | None]:
    try:
        creds = st.secrets["credenciales"]
        usuario = str(creds["usuario"]).strip()
        clave = str(creds["clave"]).strip()
        if not usuario or not clave:
            return None, None, "Faltan usuario/clave en secrets['credenciales']."
        return usuario, clave, None
    except Exception:
        return None, None, None


# Tabla única de certificados en Supabase (INSERT + SELECT)
TABLA_CERTIFICADOS = "certificados_certipe"


def get_supabase_client():
    """
    Cliente Supabase desde secrets (URL y KEY en la raíz de st.secrets).
    """
    try:
        url = str(st.secrets["SUPABASE_URL"]).strip()
        key = str(st.secrets["SUPABASE_KEY"]).strip()
    except Exception:
        # Fallback anidado (misma idea que validador)
        url = _secret_get("credenciales", "SUPABASE_URL") or _secret_get("SUPABASE_URL") or ""
        key = (
            _secret_get("credenciales", "SUPABASE_KEY")
            or _secret_get("SUPABASE_KEY")
            or ""
        )
    if not url or not key:
        raise RuntimeError(
            "Faltan SUPABASE_URL o SUPABASE_KEY en secrets "
            '(ej. SUPABASE_URL = "https://xxx.supabase.co").'
        )
    try:
        from supabase import create_client
    except ImportError as exc:
        raise RuntimeError(
            "Falta el paquete `supabase`. Añádelo a requirements.txt e instala dependencias."
        ) from exc
    return create_client(url, key)


def guardar_certificado_supabase(rec: dict) -> dict:
    """
    Inserta en public.certificados_certipe y verifica que quede guardado.
    Devuelve la fila leída desde Supabase.
    """
    client = get_supabase_client()
    codigo = (rec.get("id") or "").strip()
    if not codigo:
        raise RuntimeError("El certificado no tiene código (id).")

    fila = {
        "codigo_cert": codigo,
        "dni_alumno": rec.get("holder_doc"),
        "nombre_alumno": rec.get("holder_name"),
        "curso": rec.get("course_title"),
        "document_hash": rec.get("payload_hash"),
        "proof_value": rec.get("signature"),
        "fecha_emision": rec.get("issued_at"),
    }
    # Quitar None (evita errores de columnas opcionales)
    fila = {k: v for k, v in fila.items() if v is not None and v != ""}

    resp = (
        client.table(TABLA_CERTIFICADOS)
        .insert(fila)
        .execute()
    )
    if getattr(resp, "data", None):
        return resp.data[0]

    # Si el insert no devolvió filas, comprobar por código
    check = (
        client.table(TABLA_CERTIFICADOS)
        .select("*")
        .eq("codigo_cert", codigo)
        .limit(1)
        .execute()
    )
    if getattr(check, "data", None):
        return check.data[0]

    raise RuntimeError(
        "Supabase no devolvió la fila insertada. "
        "Revisa que existan las columnas: codigo_cert, dni_alumno, nombre_alumno, "
        "curso, document_hash, proof_value, fecha_emision. "
        f"Detalle: {resp!r}"
    )


def listar_certificados_supabase(*, limit: int | None = None) -> list[dict]:
    """SELECT exclusivo de public.certificados_certipe (más recientes primero)."""
    client = get_supabase_client()
    q = (
        client.table(TABLA_CERTIFICADOS)
        .select(
            "fecha_emision,dni_alumno,nombre_alumno,curso,document_hash,codigo_cert"
        )
        .order("fecha_emision", desc=True)
    )
    if limit is not None and limit > 0:
        q = q.limit(limit)
    resp = q.execute()
    return list(getattr(resp, "data", None) or [])


def _fmt_fecha_lista(raw: object) -> str:
    if not raw:
        return "—"
    s = str(raw)
    try:
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        return dt.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return s[:19]


def filas_alumnos_para_tabla(rows: list[dict]) -> list[dict]:
    """Columnas legibles para st.dataframe / exportación."""
    return [
        {
            "Fecha": _fmt_fecha_lista(r.get("fecha_emision")),
            "DNI": (r.get("dni_alumno") or "—"),
            "Nombre del Alumno": (r.get("nombre_alumno") or "—"),
            "Curso": (r.get("curso") or "—"),
            "Código de Verificación": (r.get("document_hash") or "—"),
        }
        for r in rows
    ]


def filtrar_alumnos(rows: list[dict], query: str) -> list[dict]:
    q = (query or "").strip().lower()
    if not q:
        return rows
    out: list[dict] = []
    for r in rows:
        dni = str(r.get("dni_alumno") or "").lower()
        nombre = str(r.get("nombre_alumno") or "").lower()
        if q in dni or q in nombre:
            out.append(r)
    return out


def exportar_lista_csv(filas: list[dict]) -> bytes:
    """
    CSV orientado a Excel (ES):
    - encoding utf-8-sig (BOM) para tildes/ñ
    - sep=';' para que Excel separe columnas automáticamente
    """
    campos = [
        "Fecha",
        "DNI",
        "Nombre del Alumno",
        "Curso",
        "Código de Verificación",
    ]
    try:
        import pandas as pd  # type: ignore

        df = pd.DataFrame(filas, columns=campos)
        bio = io.BytesIO()
        df.to_csv(
            bio,
            index=False,
            sep=";",
            encoding="utf-8-sig",
            lineterminator="\r\n",
        )
        return bio.getvalue()
    except Exception:
        # Fallback sin pandas: mismas reglas (sep=; + BOM)
        buf = io.StringIO()
        writer = csv.DictWriter(
            buf,
            fieldnames=campos,
            extrasaction="ignore",
            delimiter=";",
            lineterminator="\r\n",
            quoting=csv.QUOTE_MINIMAL,
        )
        writer.writeheader()
        writer.writerows(filas)
        return buf.getvalue().encode("utf-8-sig")


def exportar_lista_excel(filas: list[dict]) -> bytes:
    """
    Excel nativo (.xlsx) estilo packing list industrial:
    título azul oscuro, cabeceras azul medio, bordes y auto-ancho.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    # Claves del dict de filas vs títulos de columna en el archivo
    columnas = [
        ("Fecha", "Fecha"),
        ("DNI", "DNI"),
        ("Nombre del Alumno", "Alumno"),
        ("Curso", "Curso"),
        ("Código de Verificación", "Código"),
    ]
    n_cols = len(columnas)

    wb = Workbook()
    ws = wb.active
    ws.title = "Certificados"

    # Paleta packing list
    fill_title = PatternFill("solid", fgColor="1F4E79")
    fill_header = PatternFill("solid", fgColor="2E75B6")
    fill_zebra = PatternFill("solid", fgColor="F2F2F2")
    font_title = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
    font_header = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    font_cell = Font(name="Calibri", color="000000", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="B0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Fila 1: título merged
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    title_cell = ws.cell(
        row=1,
        column=1,
        value="Registro de Certificados Emitidos | CertiPE",
    )
    title_cell.fill = fill_title
    title_cell.font = font_title
    title_cell.alignment = center
    for col in range(1, n_cols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = fill_title
        c.border = border
    ws.row_dimensions[1].height = 28

    # Fila 2: cabeceras
    for col_idx, (_key, label) in enumerate(columnas, start=1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A3"

    # Datos desde fila 3
    for row_idx, fila in enumerate(filas, start=3):
        for col_idx, (key, _label) in enumerate(columnas, start=1):
            val = fila.get(key, "")
            if val is None:
                val = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=str(val))
            cell.font = font_cell
            cell.border = border
            cell.alignment = center if col_idx in (1, 2) else left
            if (row_idx - 3) % 2 == 1:
                cell.fill = fill_zebra

    # Auto-fit de columnas (aprox. openpyxl; márgenes para no cortar)
    for col_idx, (key, label) in enumerate(columnas, start=1):
        max_len = len(str(label))
        for row_idx in range(3, 3 + len(filas)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        # Límites sensatos (hash largo vs. texto legible)
        width = min(max(max_len + 3, 12), 55 if col_idx != 5 else 70)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Impresión limpia
    ws.print_title_rows = "1:2"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def app_base_url() -> str:
    explicit = (
        _secret_get("PUBLIC_BASE_URL")
        or _secret_get("credenciales", "PUBLIC_BASE_URL")
        or os.environ.get("PUBLIC_BASE_URL")
    )
    if explicit:
        return explicit.rstrip("/")
    try:
        headers = st.context.headers  # type: ignore[attr-defined]
        host = headers.get("Host") or headers.get("host")
        proto = headers.get("X-Forwarded-Proto") or headers.get("x-forwarded-proto") or "https"
        if host:
            return f"{proto}://{host}".rstrip("/")
    except Exception:
        pass
    return ""


def verify_url(cert_id: str) -> str:
    base = app_base_url()
    if base:
        return f"{base}/?codigo={quote(cert_id)}"
    return f"?codigo={quote(cert_id)}"


def inject_web_look() -> None:
    """Estilos calcados de static/styles.css (web local)."""
    st.markdown(
        """
        <style>
        @import url('https://fonts.googleapis.com/css2?family=DM+Sans:opsz,wght@9..40,400;9..40,600;9..40,700&family=IBM+Plex+Mono:wght@400;500&display=swap');

        :root {
          --ink: #14211c;
          --muted: #5b6b63;
          --ok: #1f6b45;
          --ok-bg: #e5f5eb;
          --accent: #0d6e56;
          --accent-2: #c46a1c;
          --line: #c9d5cc;
        }

        html, body, [class*="css"] {
          font-family: "DM Sans", system-ui, sans-serif !important;
          color: var(--ink);
        }

        /* Fondo como la web local */
        [data-testid="stAppViewContainer"] {
          background:
            radial-gradient(900px 500px at 10% -10%, rgba(196, 106, 28, 0.18), transparent 55%),
            radial-gradient(800px 480px at 90% 0%, rgba(13, 110, 86, 0.18), transparent 50%),
            linear-gradient(180deg, #eef4ef 0%, #f7f5f0 45%, #e7efe9 100%) !important;
        }
        [data-testid="stHeader"] { background: transparent !important; }
        [data-testid="stToolbar"] { right: 1rem; }

        .block-container {
          padding-top: 1rem !important;
          padding-bottom: 2.5rem !important;
          max-width: 920px !important;
        }

        /* Ocultar chrome de Streamlit sobrante */
        #MainMenu { visibility: hidden; }
        footer { visibility: hidden; }

        .top-nav {
          display: flex;
          justify-content: space-between;
          align-items: center;
          margin-bottom: 0.8rem;
        }
        .brand {
          font-weight: 700;
          letter-spacing: -0.03em;
          font-size: 1.25rem;
          color: var(--ink);
        }
        .nav-links { color: var(--muted); font-weight: 600; font-size: 0.95rem; }

        .eyebrow {
          text-transform: uppercase;
          letter-spacing: 0.12em;
          font-size: 0.75rem;
          font-weight: 700;
          color: var(--accent-2);
          margin: 0 0 0.4rem 0;
        }
        .hero-title {
          font-size: clamp(1.8rem, 4vw, 2.55rem);
          letter-spacing: -0.04em;
          line-height: 1.15;
          font-weight: 700;
          color: var(--ink);
          margin: 0.2rem 0 0.75rem 0;
        }
        .lead {
          color: var(--muted);
          font-size: 1.05rem;
          max-width: 38rem;
          margin: 0 0 1.1rem 0;
          line-height: 1.45;
        }

        .panel {
          background: rgba(255, 255, 255, 0.88);
          border: 1px solid rgba(201, 213, 204, 0.9);
          border-radius: 18px;
          padding: 1.25rem 1.4rem;
          box-shadow: 0 18px 50px rgba(20, 33, 28, 0.08);
          backdrop-filter: blur(8px);
          margin: 0 0 1rem 0;
        }
        .panel h2 {
          margin: 0 0 0.45rem 0;
          font-size: 1.2rem;
          color: var(--ink);
        }
        .panel .muted { color: var(--muted); font-size: 0.95rem; margin: 0 0 0.7rem 0; }

        .key-box {
          display: block;
          word-break: break-all;
          background: #e9efeb;
          padding: 0.8rem 1rem;
          border-radius: 10px;
          font-family: "IBM Plex Mono", ui-monospace, monospace;
          font-size: 0.8rem;
          color: var(--ink);
          margin-top: 0.5rem;
        }

        .list-row {
          display: flex;
          justify-content: space-between;
          align-items: center;
          gap: 0.75rem;
          padding: 0.75rem 0;
          border-bottom: 1px solid var(--line);
          font-size: 0.95rem;
        }
        .list-row:last-child { border-bottom: none; }
        .list-id {
          font-family: "IBM Plex Mono", monospace;
          font-weight: 600;
          color: var(--accent);
          font-size: 0.9rem;
        }
        .list-meta { color: var(--muted); flex: 1; }
        .tag {
          font-size: 0.72rem;
          font-weight: 700;
          text-transform: uppercase;
          letter-spacing: 0.06em;
          padding: 0.2rem 0.5rem;
          border-radius: 6px;
          background: var(--ok-bg);
          color: var(--ok);
          white-space: nowrap;
        }
        .tag.revoked { background: #f8e8e8; color: #8b2e2e; }

        .foot {
          text-align: center;
          color: var(--muted);
          font-size: 0.85rem;
          margin-top: 1.5rem;
        }

        /* Botones Streamlit → acento web */
        div.stButton > button[kind="primary"],
        div.stButton > button[data-testid="baseButton-primary"],
        div[data-testid="stFormSubmitButton"] > button,
        div[data-testid="stFormSubmitButton"] button {
          background-color: #0d6e56 !important;
          border-color: #0d6e56 !important;
          color: #fff !important;
          border-radius: 10px !important;
          font-weight: 700 !important;
        }
        div.stButton > button[kind="secondary"],
        div.stButton > button[data-testid="baseButton-secondary"] {
          background: transparent !important;
          border: 1px solid #c9d5cc !important;
          color: #14211c !important;
          border-radius: 10px !important;
          font-weight: 700 !important;
        }
        div[data-testid="stForm"] {
          background: rgba(255,255,255,0.88);
          border: 1px solid rgba(201, 213, 204, 0.9);
          border-radius: 18px;
          padding: 1rem 1.1rem 0.4rem;
          box-shadow: 0 18px 50px rgba(20, 33, 28, 0.08);
        }

        /* ── Campos de datos (sin que el texto se salga del recuadro) ── */
        [data-testid="stTextInput"] [data-baseweb="base-input"],
        [data-testid="stTextArea"] [data-baseweb="base-input"],
        [data-testid="stNumberInput"] [data-baseweb="base-input"],
        [data-testid="stTextInput"] > div > div,
        [data-testid="stTextArea"] > div > div,
        [data-testid="stNumberInput"] > div > div {
          background-color: #e9efeb !important;
          border: 1.5px solid #c9d5cc !important;
          border-radius: 10px !important;
          outline: none !important;
          box-sizing: border-box !important;
          overflow: hidden !important;
          max-width: 100% !important;
        }

        [data-testid="stTextInput"] input,
        [data-testid="stTextArea"] textarea,
        [data-testid="stNumberInput"] input,
        .stTextInput input,
        .stTextArea textarea,
        .stNumberInput input {
          background-color: transparent !important;
          color: #14211c !important;
          border: none !important;
          border-radius: 10px !important;
          min-height: 2.75rem !important;
          max-width: 100% !important;
          width: 100% !important;
          box-sizing: border-box !important;
          padding: 0.55rem 0.85rem !important;
          /* espacio al ojo (mostrar clave) y truncar placeholder largo */
          padding-right: 2.75rem !important;
          overflow: hidden !important;
          text-overflow: ellipsis !important;
          white-space: nowrap !important;
          box-shadow: none !important;
        }
        [data-testid="stTextArea"] textarea,
        .stTextArea textarea {
          min-height: 5.5rem !important;
          white-space: pre-wrap !important;
          padding-right: 0.85rem !important;
          overflow-y: auto !important;
        }

        /* Placeholder y “Press Enter…” de Streamlit: no se salen del cuadro */
        [data-testid="stTextInput"] input::placeholder,
        [data-testid="stTextArea"] textarea::placeholder,
        [data-testid="stNumberInput"] input::placeholder {
          color: #6b7a74 !important;
          opacity: 1 !important;
          overflow: hidden !important;
          text-overflow: ellipsis !important;
          white-space: nowrap !important;
        }

        /* Botón ojo de password: dentro del recuadro, no pisa el texto */
        [data-testid="stTextInput"] button,
        [data-testid="stTextInput"] [data-baseweb="input"] button,
        [data-testid="stTextInput"] button[kind="minimal"],
        [data-testid="stTextInput"] [data-testid="stTextInputShowPasswordButton"],
        [data-testid="stTextInput"] [title*="Show"],
        [data-testid="stTextInput"] [title*="Hide"] {
          flex-shrink: 0 !important;
          z-index: 2 !important;
        }

        /* Focus: un poco más claro + borde verde acento */
        [data-testid="stTextInput"]:focus-within [data-baseweb="base-input"],
        [data-testid="stTextArea"]:focus-within [data-baseweb="base-input"],
        [data-testid="stNumberInput"]:focus-within [data-baseweb="base-input"],
        [data-testid="stTextInput"]:focus-within > div > div,
        [data-testid="stTextArea"]:focus-within > div > div {
          border-color: #0d6e56 !important;
          box-shadow: 0 0 0 3px rgba(13, 110, 86, 0.18) !important;
          background-color: #e9efeb !important;
        }

        code, .stCode code {
          font-family: "IBM Plex Mono", monospace !important;
        }
        [data-testid="stMetricValue"] {
          font-family: "IBM Plex Mono", monospace;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def nav_bar(page: str) -> None:
    st.markdown(
        f"""
        <div class="top-nav">
          <div class="brand">CertiPE</div>
          <div class="nav-links">{page}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def hero_home() -> None:
    inst = certificates.institution_name()
    st.markdown(
        f"""
        <p class="eyebrow">{inst}</p>
        <h1 class="hero-title">Certificados firmados con Ed25519</h1>
        <p class="lead">
          Emite un certificado, fírmalo con la clave de la institución y compártelo.
          Cualquiera puede verificar si es auténtico o si lo alteraron.
        </p>
        """,
        unsafe_allow_html=True,
    )

def show_result_panel(result: dict, *, allow_revoke: bool = False) -> None:
    if result["ok"]:
        st.success(result["message"])
    else:
        st.error(result["message"])

    if not result.get("certificate"):
        return

    c = result["certificate"]
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Estado", c.get("status", "—"))
    m2.metric("Firma OK", str(result.get("signature_valid")))
    m3.metric("Hash OK", str(result.get("hash_valid")))
    m4.metric("Motor", result.get("verify_backend") or crypto.last_verify_backend())
    st.write(
        {
            "código": c.get("id"),
            "titular": c.get("holder_name"),
            "documento": c.get("holder_doc"),
            "curso": c.get("course_title"),
            "institución": c.get("institution_name"),
            "emitido": c.get("issued_at"),
            "hash": c.get("payload_hash"),
        }
    )
    link = verify_url(c["id"])
    st.markdown("**Link de verificación**")
    st.code(link, language=None)
    try:
        st.image(pdf_cert.qr_png_bytes(link), caption="QR de validación", width=180)
    except Exception:
        pass
    try:
        pdf_bytes = pdf_cert.build_certificate_pdf(c, link)
        st.download_button(
            "Descargar PDF",
            data=pdf_bytes,
            file_name=f"{c['id']}.pdf",
            mime="application/pdf",
            key=f"pdf_{c['id']}_{allow_revoke}",
        )
    except Exception as e:  # noqa: BLE001
        st.warning(f"PDF no disponible: {e}")

    if allow_revoke and c.get("status") == "valid":
        if st.button("Revocar este certificado", type="secondary", key=f"rev_{c['id']}"):
            certificates.revoke(c["id"])
            st.warning("Revocado.")
            st.rerun()


def panel_pubkey() -> None:
    fp = crypto.public_key_fingerprint()
    pub = crypto.public_key_b64()
    st.markdown(
        f"""
        <div class="panel">
          <h2>Clave pública de la institución</h2>
          <p class="muted">Huella: <code>{fp}</code>. La privada nunca sale del servidor.</p>
          <div class="key-box">{pub}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def _esc_html(text: object) -> str:
    s = str(text if text is not None else "")
    return (
        s.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def panel_recent() -> None:
    """Últimos emitidos: solo public.certificados_certipe (no JSON local ni otras tablas)."""
    rows = ""
    try:
        items = listar_certificados_supabase(limit=12)
    except Exception as e:  # noqa: BLE001
        st.markdown(
            f"""
            <div class="panel">
              <h2>Últimos emitidos</h2>
              <div class="list-row">
                <span class="list-meta">No se pudo leer Supabase ({_esc_html(e)}).</span>
              </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        return

    if not items:
        rows = (
            '<div class="list-row">'
            '<span class="list-meta">Aún no hay certificados en la base de datos.</span>'
            "</div>"
        )
    else:
        for c in items:
            codigo = _esc_html(c.get("codigo_cert") or "—")
            nombre = _esc_html(c.get("nombre_alumno") or "—")
            curso = _esc_html(c.get("curso") or "—")
            rows += (
                f'<div class="list-row">'
                f'<span class="list-id">{codigo}</span>'
                f'<span class="list-meta">{nombre} · {curso}</span>'
                f'<span class="tag">VALID</span>'
                f"</div>"
            )

    st.markdown(
        f"""
        <div class="panel">
          <h2>Últimos emitidos</h2>
          {rows}
        </div>
        """,
        unsafe_allow_html=True,
    )


def foot() -> None:
    st.markdown(
        f'<p class="foot">MVP · Ed25519 · {APP_VERSION} · demo institucional</p>',
        unsafe_allow_html=True,
    )


# ── bootstrap ───────────────────────────────────────────────────────────────
bootstrap_secrets()
crypto.ensure_keys()
inject_web_look()

if "autenticado" not in st.session_state:
    st.session_state["autenticado"] = False
if "page" not in st.session_state:
    # Home público = consulta / validar (clave pública + últimos emitidos)
    st.session_state["page"] = "inicio"

qp = st.query_params
codigo_qp = (qp.get("codigo") or qp.get("v") or "").strip().upper()
page_qp = (qp.get("page") or "").strip().lower()
mode_qp = (qp.get("public") or qp.get("mode") or "").strip().lower()

# Deep links por URL
if mode_qp in {"1", "true", "validar", "validate"} or page_qp in {
    "validar",
    "validar_publico",
}:
    st.session_state["page"] = "validar_publico"
elif page_qp in {"inicio", "home"}:
    st.session_state["page"] = "inicio"
elif page_qp == "emitir":
    st.session_state["page"] = "emitir"
elif page_qp == "lista":
    st.session_state["page"] = "lista"
elif page_qp == "crypto":
    st.session_state["page"] = "crypto"

# Validación profunda por link (sin login)
if codigo_qp:
    nav_bar("Validar")
    st.markdown(
        f"""
        <p class="eyebrow">{certificates.institution_name()}</p>
        <h1 class="hero-title">Verificación pública</h1>
        <p class="lead">Código <code>{codigo_qp}</code></p>
        """,
        unsafe_allow_html=True,
    )
    with st.container():
        show_result_panel(certificates.validate(codigo_qp), allow_revoke=False)
    foot()
    st.stop()

usuario_cfg, clave_cfg, err_creds = cargar_credenciales()
login_requerido = usuario_cfg is not None and clave_cfg is not None

# Login solo para emitir / lista / admin — validar y ver inicio pueden ser mixtos
need_login_pages = {"emitir", "lista", "crypto"}
page = st.session_state["page"]

if (
    login_requerido
    and not st.session_state["autenticado"]
    and page in need_login_pages
):
    nav_bar("Acceso")
    st.markdown(
        f"""
        <p class="eyebrow">{certificates.institution_name()}</p>
        <h1 class="hero-title">Acceso institución</h1>
        <p class="lead">Inicia sesión para emitir certificados. La validación pública no pide clave.</p>
        """,
        unsafe_allow_html=True,
    )
    b1, b2 = st.columns(2)
    with b1:
        if st.button("Validar código", use_container_width=True, type="secondary"):
            st.session_state["page"] = "validar_publico"
            st.rerun()
    with b2:
        if st.button("Volver al inicio", use_container_width=True, type="secondary"):
            st.session_state["page"] = "inicio"
            st.rerun()

    if err_creds:
        st.error(err_creds)

    # st.form nativo (sin divs HTML rotos) → Enter en Usuario/Clave envía "Entrar"
    with st.form("login", clear_on_submit=False, border=False):
        st.markdown("##### Acceso")
        u = st.text_input("Usuario", key="login_usuario", autocomplete="username")
        p = st.text_input(
            "Clave",
            type="password",
            key="login_clave",
            autocomplete="current-password",
        )
        entrar = st.form_submit_button(
            "Entrar",
            type="primary",
            use_container_width=True,
        )

    if entrar:
        if u.strip() == usuario_cfg and p == clave_cfg:
            st.session_state["autenticado"] = True
            st.session_state["page"] = "emitir"
            st.rerun()
        st.error("Usuario o clave incorrectos.")
    foot()
    st.stop()

# ── Páginas ─────────────────────────────────────────────────────────────────
if page == "inicio":
    nav_bar("Inicio")
    hero_home()
    c1, c2 = st.columns(2)
    with c1:
        if st.button("Emitir certificado", type="primary", use_container_width=True):
            st.session_state["page"] = "emitir"
            st.rerun()
    with c2:
        if st.button("Validar código", type="secondary", use_container_width=True):
            st.session_state["page"] = "validar_publico"
            st.rerun()

    if "seed_error" in st.session_state:
        st.error(f"Error en LLAVE_PRIVADA: {st.session_state['seed_error']}")

    panel_pubkey()
    panel_recent()

    if login_requerido and st.session_state["autenticado"]:
        a1, a2, a3 = st.columns(3)
        with a1:
            if st.button("Lista", use_container_width=True):
                st.session_state["page"] = "lista"
                st.rerun()
        with a2:
            if st.button("Criptografía", use_container_width=True):
                st.session_state["page"] = "crypto"
                st.rerun()
        with a3:
            if st.button("Cerrar sesión", use_container_width=True):
                st.session_state["autenticado"] = False
                st.rerun()
    elif login_requerido:
        if st.button("Entrar como institución", type="secondary", use_container_width=True):
            st.session_state["page"] = "emitir"
            st.rerun()

    foot()

elif page in {"validar", "validar_publico"}:
    nav_bar("Validar")
    st.markdown(
        f"""
        <p class="eyebrow">{certificates.institution_name()}</p>
        <h1 class="hero-title">Validar certificado</h1>
        <p class="lead">Ingresa el código (ej. CERT-XXXX-XXXX-XXXX) y pulsa Enter o Verificar.</p>
        """,
        unsafe_allow_html=True,
    )
    with st.form("validar_codigo", clear_on_submit=False, border=False):
        codigo = st.text_input(
            "Código",
            placeholder="CERT-....",
            label_visibility="collapsed",
            key="validar_codigo_input",
        )
        go = st.form_submit_button(
            "Verificar firma",
            type="primary",
            use_container_width=True,
        )
    if st.button("← Inicio", type="secondary", use_container_width=True):
        st.session_state["page"] = "inicio"
        st.rerun()
    if go and codigo.strip():
        show_result_panel(
            certificates.validate(codigo.strip().upper()),
            allow_revoke=bool(st.session_state.get("autenticado")),
        )
    foot()

elif page == "emitir":
    nav_bar("Emitir")
    st.markdown(
        f"""
        <p class="eyebrow">{certificates.institution_name()}</p>
        <h1 class="hero-title">Emitir certificado</h1>
        <p class="lead">La institución firmará el documento con Ed25519.</p>
        """,
        unsafe_allow_html=True,
    )
    # Tras un emit exitoso se incrementa el contador → form nuevo vacío
    form_id = st.session_state.get("emit_form_id", 0)
    with st.form(f"emitir_{form_id}", clear_on_submit=True):
        holder_name = st.text_input(
            "Nombre del titulado",
            placeholder="Denilson Aurely Padilla Arevalo",
        )
        holder_doc = st.text_input("Documento (DNI / CE)", placeholder="12345678")
        course_title = st.text_input("Curso / programa", placeholder="Power Bi desde cero")
        c1, c2 = st.columns(2)
        with c1:
            course_hours = st.number_input("Horas", min_value=0, max_value=10000, value=0)
            grade = st.text_input("Nota (opcional)", placeholder="16 (Dieciséis)")
            city = st.text_input("Ciudad", placeholder="Trujillo")
        with c2:
            issued_by = st.text_input("Nombre del firmante", placeholder="Director académico")
            signer_role = st.text_input("Cargo del firmante", placeholder="Director Académico Nacional")
            notes = st.text_input(
                "Periodo / detalle",
                placeholder="Desarrollado del 14 de mayo de 2026 al 9 de junio de 2026",
            )
        submitted = st.form_submit_button("Firmar y emitir", type="primary", use_container_width=True)
        if submitted:
            if not holder_name.strip() or not holder_doc.strip() or not course_title.strip():
                st.warning("Completa nombre, documento y curso.")
            else:
                rec = certificates.issue(
                    holder_name=holder_name,
                    holder_doc=holder_doc,
                    course_title=course_title,
                    course_hours=int(course_hours) or None,
                    issued_by=issued_by or None,
                    notes=notes or None,
                    grade=grade or None,
                    city=city or None,
                    signer_role=signer_role or None,
                )
                st.session_state["last_cert"] = rec
                st.session_state["supabase_ok"] = False
                st.session_state["supabase_error"] = None
                try:
                    fila_sb = guardar_certificado_supabase(rec)
                    st.session_state["supabase_ok"] = True
                    st.session_state["supabase_row"] = fila_sb
                except Exception as e:  # noqa: BLE001
                    st.session_state["supabase_error"] = str(e)
                    st.session_state["supabase_row"] = None
                # Vaciar formulario para el siguiente certificado
                st.session_state["emit_form_id"] = form_id + 1
                st.rerun()

    if st.button("← Inicio", type="secondary"):
        st.session_state["page"] = "inicio"
        st.rerun()

    if st.session_state.get("last_cert"):
        rec = st.session_state["last_cert"]
        st.success(f"Emitido: **{rec['id']}**")
        if st.session_state.get("supabase_ok"):
            st.success(
                "✅ Certificado guardado de forma segura en la base de datos de Supabase"
            )
            fila = st.session_state.get("supabase_row") or {}
            if fila:
                st.caption(
                    f"Supabase · codigo_cert=`{fila.get('codigo_cert') or '—'}` · "
                    f"nombre=`{fila.get('nombre_alumno') or '—'}` · "
                    f"curso=`{fila.get('curso') or '—'}`"
                )
        elif st.session_state.get("supabase_error"):
            st.error(f"No se pudo guardar en Supabase: {st.session_state['supabase_error']}")
        link = verify_url(rec["id"])
        st.code(link, language=None)
        ca, cb = st.columns([1, 2])
        with ca:
            try:
                st.image(pdf_cert.qr_png_bytes(link), width=160)
            except Exception as e:  # noqa: BLE001
                st.caption(str(e))
        with cb:
            try:
                pdf_bytes = pdf_cert.build_certificate_pdf(rec, link)
                st.download_button(
                    "Descargar PDF del certificado",
                    data=pdf_bytes,
                    file_name=f"{rec['id']}.pdf",
                    mime="application/pdf",
                    type="primary",
                    key="pdf_last",
                )
            except Exception as e:  # noqa: BLE001
                st.warning(f"PDF: {e}")
    foot()

elif page == "lista":
    nav_bar("Lista")
    st.markdown('<h1 class="hero-title">Lista de alumnos</h1>', unsafe_allow_html=True)
    st.caption(f"Registros en Supabase · `public.{TABLA_CERTIFICADOS}`")
    if st.button("← Inicio", type="secondary"):
        st.session_state["page"] = "inicio"
        st.rerun()

    busqueda = st.text_input(
        "Buscar por nombre o DNI",
        placeholder="Ej. Padilla o 12345678",
        key="lista_busqueda",
    )

    try:
        registros = listar_certificados_supabase()
    except Exception as e:  # noqa: BLE001
        st.error(f"No se pudo consultar Supabase: {e}")
        st.info(
            "Comprueba `SUPABASE_URL`, `SUPABASE_KEY` y que la tabla "
            f"`public.{TABLA_CERTIFICADOS}` exista y permita SELECT."
        )
        foot()
        st.stop()

    filtrados = filtrar_alumnos(registros, busqueda)
    filas = filas_alumnos_para_tabla(filtrados)

    st.markdown(
        f"**{len(filas)}** registro(s)"
        + (f" (de {len(registros)} totales)" if busqueda.strip() else "")
    )

    st.dataframe(
        filas,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Código de Verificación": st.column_config.TextColumn(
                "Código de Verificación",
                width="large",
            ),
            "Nombre del Alumno": st.column_config.TextColumn(
                "Nombre del Alumno",
                width="medium",
            ),
        },
    )

    if not filas:
        st.info("No hay alumnos que coincidan con la búsqueda (o la tabla está vacía).")
    else:
        stamp = datetime.now().strftime("%Y%m%d_%H%M")
        col_csv, col_xlsx = st.columns(2)
        with col_csv:
            st.download_button(
                "Descargar CSV",
                data=exportar_lista_csv(filas),
                file_name=f"lista_alumnos_certipe_{stamp}.csv",
                mime="text/csv",
                use_container_width=True,
                key="dl_lista_csv",
            )
        with col_xlsx:
            st.download_button(
                "Descargar Excel",
                data=exportar_lista_excel(filas),
                file_name=f"lista_alumnos_certipe_{stamp}.xlsx",
                mime=(
                    "application/vnd.openxmlformats-officedocument"
                    ".spreadsheetml.sheet"
                ),
                use_container_width=True,
                key="dl_lista_xlsx",
            )

    foot()

elif page == "crypto":
    nav_bar("Cripto")
    st.markdown('<h1 class="hero-title">Criptografía</h1>', unsafe_allow_html=True)
    if st.button("← Inicio", type="secondary"):
        st.session_state["page"] = "inicio"
        st.rerun()
    panel_pubkey()
    st.write(
        {
            "alg": "Ed25519",
            "fingerprint": crypto.public_key_fingerprint(),
            "public_key_hex": crypto.public_key_hex(),
            "motor_rust": crypto.motor_rust_disponible(),
            "last_verify_backend": crypto.last_verify_backend(),
            "app_version": APP_VERSION,
        }
    )
    foot()
